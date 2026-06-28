"""Lectura de archivos Excel (.xlsx) y tablas HTML en el pipeline de import.

Algunos brokers exportan en formatos que no son CSV plano:
  • Bull Market → Excel binario (.xlsx).
  • IOL (InvertirOnline) → "Movimientos Históricos" en un .xls que en realidad
    es una tabla HTML (un <table> generado por el sitio). No es Excel binario
    ni BIFF: si lo abrís con un editor de texto ves HTML.

En vez de escribir parsers que entiendan binario o HTML, convertimos todo a
texto CSV acá, y los parsers existentes (text-based) lo procesan igual que un
CSV.

Diseño:
  • is_xlsx(bytes): detección por magic bytes (xlsx = zip → 'PK\\x03\\x04').
    Robusto: no depende del filename.
  • xlsx_to_csv(bytes): primera hoja → CSV (coma-separated). Fechas a ISO
    (YYYY-MM-DD) para que los parsers no tengan que adivinar formato; números
    tal cual los da openpyxl (point-decimal). Filas 100% vacías se descartan.
  • is_html_table(bytes) / html_table_to_csv(bytes): detecta y aplana la PRIMERA
    tabla HTML a CSV usando solo la stdlib (html.parser) — sin pandas/bs4/lxml,
    que NO están en requirements.txt y faltarían en prod.
"""
from __future__ import annotations

import csv
import io
from datetime import datetime, date
from html.parser import HTMLParser


def is_xlsx(file_bytes: bytes) -> bool:
    """True si los bytes son un .xlsx (archivo zip con firma PK\\x03\\x04).
    Los .xls viejos (BIFF) NO matchean — solo soportamos xlsx."""
    return bool(file_bytes) and file_bytes[:4] == b"PK\x03\x04"


def is_pdf(file_bytes: bytes) -> bool:
    """True si los bytes son un PDF (firma '%PDF-')."""
    return bool(file_bytes) and file_bytes[:5] == b"%PDF-"


def pdf_to_text(file_bytes: bytes) -> str:
    """Extrae el TEXTO (layout-aware) de un PDF con pdfplumber. Lo usa el flujo de
    "Tenencia valorizada" de Bull Market — la imprime en PDF, sin export a Excel.
    Import perezoso para no romper la carga del módulo si pdfplumber faltara en
    algún entorno (falla recién al intentar leer un PDF)."""
    try:
        import pdfplumber
    except ImportError:
        raise ValueError(
            "No pudimos leer el PDF (falta el lector en el servidor). Escribinos "
            "y lo habilitamos — o subí el reporte en Excel/CSV si lo tenés.")
    try:
        parts = []
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            for page in pdf.pages:
                parts.append(page.extract_text() or "")
        return "\n".join(parts)
    except Exception as ex:
        raise ValueError(f"No pudimos abrir el PDF: {ex}")


def _cell_to_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        # ISO date — los parsers la leen sin ambigüedad de formato.
        return v.strftime("%Y-%m-%d")
    if isinstance(v, float):
        # Evita notación científica para números grandes; point-decimal.
        return repr(v)
    return str(v).strip()


def _sheet_rows(ws) -> "list[list[str]]":
    """Filas NO vacías de una hoja, como listas de strings. La 1ª es el header."""
    out = []
    for row in ws.iter_rows(values_only=True):
        cells = [_cell_to_str(v) for v in row]
        if any(c.strip() for c in cells):  # descartar separadores / footers vacíos
            out.append(cells)
    return out


def xlsx_to_csv(file_bytes: bytes) -> str:
    """Convierte un .xlsx a texto CSV (coma-separated).

    - 1 hoja  → header + filas, con una columna sintética '_hoja' = título de la
      hoja (la usan parsers que infieren moneda por hoja).
    - N hojas → UNIÓN de columnas de todas las hojas (por nombre de header,
      preservando orden de aparición) + '_hoja' = título de la hoja de cada fila.
      Así un export multi-reporte (ej. Resultados de Balanz: lotes_iniciales /
      por_realizado / lotes_finales) llega entero al parser, que dispatcha por
      '_hoja'. Filas de una hoja sin una columna de la unión van vacías.

    Levanta ValueError con mensaje claro si el archivo no se puede abrir."""
    try:
        import openpyxl
    except ImportError as ex:  # pragma: no cover
        raise ValueError("Falta la librería openpyxl para leer Excel.") from ex

    try:
        # read_only=False a propósito: algunos exports (ej. Bull Market cuenta
        # cable) graban mal la dimensión de la hoja y en modo read_only openpyxl
        # trunca columnas. El modo normal lee la grilla real (ok para ≤5 MB).
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as ex:
        raise ValueError(
            "No pudimos abrir el Excel. Verificá que sea un .xlsx válido."
        ) from ex

    try:
        if not wb.worksheets:
            raise ValueError("El Excel no tiene hojas.")
        # (título, filas) por hoja con datos.
        sheets = [((ws.title or ""), _sheet_rows(ws)) for ws in wb.worksheets]
        sheets = [(t, r) for t, r in sheets if r]
        if not sheets:
            raise ValueError("El Excel no tiene datos.")

        # Unión de columnas (por nombre de header) preservando orden de aparición.
        union: list[str] = []
        seen: set = set()
        for _title, rows in sheets:
            for h in rows[0]:
                key = h.strip()
                if key and key not in seen:
                    seen.add(key)
                    union.append(h)

        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(union + ["_hoja"])
        for title, rows in sheets:
            hdr_pos = {h.strip(): i for i, h in enumerate(rows[0])}
            for r in rows[1:]:
                mapped = []
                for col in union:
                    i = hdr_pos.get(col.strip())
                    mapped.append(r[i] if (i is not None and i < len(r)) else "")
                writer.writerow(mapped + [title])
        return buf.getvalue()
    finally:
        wb.close()


def xlsx_to_rows(file_bytes: bytes, sheet_index: int = 0) -> "list[list]":
    """Filas CRUDAS (valores sin convertir) de UNA hoja del .xlsx.

    A diferencia de `xlsx_to_csv`, NO asume que la fila 0 es el header, NO une
    hojas y NO stringifica: devuelve los valores tal cual los da openpyxl
    (None / str / int / float / datetime). Lo usan parsers cuya grilla tiene
    preámbulo + secciones (ej. el "Estado de Cuenta" de PPI), donde tratar la
    fila 0 como header rompería la estructura.

    Levanta ValueError con mensaje claro si el archivo no se puede abrir."""
    try:
        import openpyxl
    except ImportError as ex:  # pragma: no cover
        raise ValueError("Falta la librería openpyxl para leer Excel.") from ex
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as ex:
        raise ValueError(
            "No pudimos abrir el Excel. Verificá que sea un .xlsx válido."
        ) from ex
    try:
        if not wb.worksheets:
            raise ValueError("El Excel no tiene hojas.")
        idx = sheet_index if 0 <= sheet_index < len(wb.worksheets) else 0
        return [list(r) for r in wb.worksheets[idx].iter_rows(values_only=True)]
    finally:
        wb.close()


# ─── Tablas HTML (.xls de IOL y similares) ───────────────────────────────────

def _decode_html_bytes(file_bytes: bytes) -> str:
    """Decodifica bytes HTML a texto probando encodings comunes. El export de
    IOL es UTF-8; cp1252/latin-1 cubren variantes de Windows. Último recurso:
    latin-1 con replace (nunca tira)."""
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return file_bytes.decode(enc)
        except UnicodeDecodeError:
            continue
    return file_bytes.decode("latin-1", "replace")


def is_html_table(file_bytes: bytes) -> bool:
    """True si los bytes parecen un documento HTML con una tabla. Cubre el .xls
    de IOL (que es un <table> HTML, no Excel binario).

    Conservador en dos frentes para no robarle archivos al path de CSV:
    1. El contenido debe ARRANCAR con '<' (tras BOM/whitespace). Así un CSV que
       trae '<table>' adentro de una celda (ej. una nota) NO se enruta acá.
    2. Además exige '<table' y al menos una fila/celda en el prefijo.
    """
    if not file_bytes:
        return False
    head = file_bytes[:16384]
    stripped = head.lstrip()
    if stripped[:3] == b"\xef\xbb\xbf":   # BOM UTF-8
        stripped = stripped[3:].lstrip()
    if not stripped.startswith(b"<"):     # un CSV real no empieza con '<'
        return False
    low = head.lower()
    return b"<table" in low and (b"<tr" in low or b"<td" in low)


class _FirstTableExtractor(HTMLParser):
    """Extrae las filas de la PRIMERA <table> del HTML como listas de strings.
    Acumula el texto de cada <td>/<th> (colapsando whitespace) y corta apenas
    cierra la primera tabla — exports de un solo cuadro como el de IOL."""

    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)  # &aacute; etc. → texto
        self.rows: list[list[str]] = []
        self._table_depth = 0
        self._done = False
        self._row: list[str] | None = None
        self._cell: list[str] | None = None

    def handle_starttag(self, tag, attrs):
        if self._done:
            return
        t = tag.lower()
        if t == "table":
            self._table_depth += 1
        elif t == "tr" and self._table_depth == 1:
            # Solo filas de la tabla EXTERIOR. Un <tr> de una tabla anidada
            # (depth>1) no debe pisar la fila en curso ni crear filas espurias.
            self._row = []
        elif t in ("td", "th") and self._table_depth == 1 and self._row is not None:
            # Solo celdas del nivel exterior. Las de una tabla anidada se
            # ignoran (no arrancan _cell), así no se cuelan en la fila exterior.
            self._cell = []

    def handle_data(self, data):
        # Solo texto de celdas del nivel exterior (depth 1). El de tablas
        # anidadas se descarta — no debe colarse en la celda en curso.
        if self._cell is not None and self._table_depth == 1:
            self._cell.append(data)

    def handle_endtag(self, tag):
        if self._done:
            return
        t = tag.lower()
        if t == "table":
            # El cierre de tabla SIEMPRE decrementa (puede ser una anidada).
            if self._table_depth:
                self._table_depth -= 1
                if self._table_depth == 0:
                    self._done = True  # solo la primera tabla (la exterior)
            return
        if self._table_depth != 1:
            return  # td/th/tr de tablas anidadas → ignorar
        if t in ("td", "th") and self._cell is not None and self._row is not None:
            self._row.append(" ".join("".join(self._cell).split()))
            self._cell = None
        elif t == "tr" and self._row is not None:
            self.rows.append(self._row)
            self._row = None


def html_table_to_csv(file_bytes: bytes) -> str:
    """Convierte la PRIMERA tabla HTML del archivo a texto CSV (coma-separated).

    Levanta ValueError con mensaje claro si no hay una tabla con datos (el
    caller lo traduce a un error de import amigable)."""
    text = _decode_html_bytes(file_bytes)
    extractor = _FirstTableExtractor()
    try:
        extractor.feed(text)
        extractor.close()
    except Exception as ex:
        raise ValueError(
            "No pudimos leer la tabla del archivo. Verificá que sea el export de movimientos."
        ) from ex
    # Si quedó una fila/celda abierta sin su </tr>/</td>, el archivo se cortó a
    # la mitad (descarga interrumpida) — html.parser.close() no flushea tags
    # abiertos, así que esa última fila se perdería en silencio. Avisamos.
    if extractor._row is not None or extractor._cell is not None:
        raise ValueError(
            "El archivo parece incompleto (la descarga se cortó). Volvé a descargar el export."
        )
    rows = [r for r in extractor.rows if any(c.strip() for c in r)]
    if not rows:
        raise ValueError("El archivo no contiene una tabla con datos.")
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerows(rows)
    return buf.getvalue()


def to_csv_text(file_bytes: bytes) -> str:
    """Devuelve texto CSV a partir de bytes que pueden ser .xlsx, una tabla
    HTML (.xls de IOL) o un CSV ya en texto. Único punto de entrada para el
    pipeline: 'dame esto como CSV'.

    - xlsx       → convierte la primera hoja.
    - tabla HTML → aplana la primera <table> (IOL).
    - CSV        → decodifica (utf-8-sig, fallback latin-1).
    Levanta ValueError si no se puede interpretar.
    """
    if is_xlsx(file_bytes):
        return xlsx_to_csv(file_bytes)
    if is_html_table(file_bytes):
        return html_table_to_csv(file_bytes)
    for enc in ("utf-8-sig", "latin-1"):
        try:
            return file_bytes.decode(enc)
        except UnicodeDecodeError:
            continue
    raise ValueError("No pudimos decodificar el archivo. Probá guardarlo como UTF-8 o .xlsx.")
