"""E2E del Portafolio de IEB aplicado en modo OVERRIDE (mismo mecanismo que Balanz:
ventas transfer_out reversibles + guardas), vía el endpoint HTTP real.

Cubre lo IEB-específico del wiring:
  • detección is_ieb + parse del Portafolio (xlsx multi-hoja) por el endpoint;
  • override que REDUCE el excedente (venta a costo, P&L 0) cuando NO se capea;
  • el CAP de sanidad frena un ×3 de cartera entera (no auto-vacía) → sólo reporta;
  • `complete` gateado por completitud: una foto PARCIAL (warnings) NO borra
    not_in_snapshot (no destruye una posición todavía tenida).

Corre con: cd backend && python3 -m pytest tests/test_ieb_tenencia_override_e2e.py
"""
import io
import os
import sys
import tempfile
import unittest
from datetime import datetime

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
BACKEND = os.path.dirname(HERE)
if BACKEND not in sys.path:
    sys.path.insert(0, BACKEND)

_TMP = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
_TMP.close()
os.environ["DB_PATH"] = _TMP.name

from importing import pipeline as pl        # noqa: E402
from importing import persister as ps        # noqa: E402
from importing import rebuild as rb          # noqa: E402
from importing.schema import NormalizedTx, OP_BUY, OP_DEPOSIT  # noqa: E402
import main                                  # noqa: E402
from fastapi.testclient import TestClient    # noqa: E402

HDR = ["Especie", "Moneda de emisión", "Cantidad", "Precio", "% del total",
       "PPP", "Var%", "Resultado", "Actualizado", "Posición total"]


def _helpers():
    h = main._ImportHelpers()
    for n in ("_adjust_broker_cash", "_adjust_cash", "_update_monthly_pnl_realized",
              "_update_monthly_flow", "_repair_monthly_chain", "_ensure_usd_sibling",
              "_recalc_pnl_realized_from_ops"):
        setattr(h, n, getattr(main, n))
    return h


def _row(tkr, qty, ppp, sect_name):
    return [f"{tkr} - {tkr}", "ARS", str(qty), str(ppp), "5.0",
            str(ppp), "1.0", "0", "10:00", str(qty * ppp)]


def _ieb_wb(sections, saldos=("1000", "0"), extra_sheet=None):
    """sections: list of (section_name, [(tkr, qty, ppp)])."""
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Patrimonio"
    ws.append(["Fecha:", datetime(2026, 7, 1, 3, 0, 0)])
    ws.append(["Patrimonio total"])
    for name, holdings in sections:
        ws.append([name]); ws.append(list(HDR))
        for tkr, qty, ppp in holdings:
            ws.append(_row(tkr, qty, ppp, name))
        ws.append(["Subtotal", "-", "-", None, "-", "-", "0", "-", "0", None])
    sal = wb.create_sheet("Saldos")
    for marker, total in (("ARS", saldos[0]), ("USD", saldos[1])):
        sal.append([marker]); sal.append(["Plazo", "Fecha", "Saldo"])
        sal.append(["Hoy", "2026-07-01", total]); sal.append(["Total", "-", total])
        sal.append([])
    if extra_sheet:
        ex = wb.create_sheet(extra_sheet)
        ex.append(["Especie", "Moneda", "Cantidad", "Tasa"])
        ex.append(["CAUCION 7D", "ARS", "500000", "40.0"])
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


class IebOverrideE2E(unittest.TestCase):
    BROKER = "IEB"

    def setUp(self):
        self.conn = main.get_db()
        for t in ("import_op_links", "import_normalized_tx", "import_raw_rows",
                  "import_batches", "operations", "positions", "monthly_entries",
                  "snapshots", "config", "brokers", "users"):
            try:
                self.conn.execute(f"DELETE FROM {t}")
            except Exception:
                pass
        self.uid = self.conn.execute(
            "INSERT INTO users (email, password_hash, approved) VALUES (?,?,1)",
            ("ieb_ovr@rendi.test", "x")).lastrowid
        self.conn.execute("INSERT INTO brokers (user_id, name, currency) VALUES (?,?,?)",
                          (self.uid, self.BROKER, "ARS"))
        self.conn.execute("INSERT OR REPLACE INTO config (user_id, key, value) VALUES (?,?,?)",
                          (self.uid, "tc_blue", "1000"))
        self.conn.commit()
        self.token = main.create_token(self.uid)
        self.client = TestClient(main.app)

    def tearDown(self):
        self.conn.close()

    def _buy(self, tkr, qty, price, at="CEDEAR", date="2026-01-10"):
        return NormalizedTx(row_index=0, date=date, broker=self.BROKER, operation_type=OP_BUY,
                            asset_symbol=tkr, asset_type=at, quantity=qty, unit_price=price,
                            gross_amount=qty * price, currency="ARS")

    def _import_mov(self, txs):
        # DEPOSITO que financia + BUYs, import-linked (→ _is_safe_to_rebuild True).
        dep = NormalizedTx(row_index=0, date="2026-01-09", broker=self.BROKER,
                           operation_type=OP_DEPOSIT,
                           gross_amount=sum(t.gross_amount for t in txs) + 100000, currency="ARS")
        allt = [dep] + txs
        with self.conn:
            for i, t in enumerate(allt):
                t.row_index = -100 - i
            sid = pl.store_preview_txs(self.conn, self.uid, broker=self.BROKER,
                                       parser_format="ieb", file_name="mov.xlsx", txs=allt)
            txs2, raw = pl.load_session_for_confirm(self.conn, uid=self.uid, session_id=sid)
            ps.persist_batch(self.conn, uid=self.uid, batch_id=sid, txs=txs2,
                             raw_row_ids_by_index=raw, helpers=_helpers())
            tc = ps._read_tc_blue(self.conn, uid=self.uid)
            rb.rebuild_fifo_after_import(self.conn, self.uid, sid, tc_blue=tc)

    def _import_mov_broker(self, broker, ccy, txs):
        # Importa BUYs (import-linked) en un broker específico (ej. el sibling USD).
        dep = NormalizedTx(row_index=0, date="2026-01-09", broker=broker,
                           operation_type=OP_DEPOSIT,
                           gross_amount=sum(t.gross_amount for t in txs) + 1000, currency=ccy)
        allt = [dep] + txs
        with self.conn:
            for i, t in enumerate(allt):
                t.row_index = -300 - i
            sid = pl.store_preview_txs(self.conn, self.uid, broker=broker,
                                       parser_format="ieb", file_name="movu.xlsx", txs=allt)
            txs2, raw = pl.load_session_for_confirm(self.conn, uid=self.uid, session_id=sid)
            ps.persist_batch(self.conn, uid=self.uid, batch_id=sid, txs=txs2,
                             raw_row_ids_by_index=raw, helpers=_helpers())
            rb.rebuild_fifo_after_import(self.conn, self.uid, sid,
                                         tc_blue=ps._read_tc_blue(self.conn, uid=self.uid))

    def _held(self, asset):
        r = self.conn.execute("SELECT COALESCE(SUM(quantity),0) q FROM positions "
                             "WHERE user_id=? AND asset=? AND is_cash=0", (self.uid, asset)).fetchone()
        return float(r["q"] or 0)

    def _held_in(self, broker, asset):
        r = self.conn.execute("SELECT COALESCE(SUM(quantity),0) q FROM positions "
                             "WHERE user_id=? AND broker=? AND asset=? AND is_cash=0",
                             (self.uid, broker, asset)).fetchone()
        return float(r["q"] or 0)

    def test_override_reduces_usd_sibling_holding(self):
        """La foto PISA los holdings en DÓLARES (sibling USD), no sólo el padre ARS:
        GLOB 100→60 en el sibling contra la foto USD (con AAPL/NVDA match para no
        capear). Antes el path agregado saltaba el sibling (same-broker guard)."""
        SIB = self.BROKER + " · USD"
        self.conn.execute("INSERT INTO brokers (user_id, name, currency) VALUES (?,?,?)",
                          (self.uid, SIB, "USDT"))
        self.conn.commit()
        self._import_mov([self._buy("MELI", 52, 21112)])          # ARS en el padre
        self._import_mov_broker(SIB, "USD", [                      # USD en el sibling
            self._usd_buy(SIB, "GLOB", 100, 50), self._usd_buy(SIB, "AAPL", 10, 200),
            self._usd_buy(SIB, "NVDA", 5, 120)])
        self.assertAlmostEqual(self._held_in(SIB, "GLOB"), 100)

        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Patrimonio"
        ws.append(["Fecha:", datetime(2026, 7, 1, 3, 0, 0)]); ws.append(["Patrimonio total"])
        ws.append(["Cedears"]); ws.append(list(HDR))
        ws.append(_row("MELI", 52, 21112, "Cedears"))
        for tkr, qty, ppp in (("GLOB", 60, 50), ("AAPL", 10, 200), ("NVDA", 5, 120)):
            ws.append([f"{tkr} - {tkr}", "USD", str(qty), str(ppp), "5.0",
                       str(ppp), "1.0", "0", "10:00", str(qty * ppp)])
        ws.append(["Subtotal", "-", "-", None, "-", "-", "0", "-", "0", None])
        sal = wb.create_sheet("Saldos")
        for m, t in (("ARS", "1000"), ("USD", "0")):
            sal.append([m]); sal.append(["Plazo", "Fecha", "Saldo"])
            sal.append(["Hoy", "2026-07-01", t]); sal.append(["Total", "-", t]); sal.append([])
        buf = io.BytesIO(); wb.save(buf)
        body = self._preview(buf.getvalue()).json()
        self.assertTrue(body["foto_completa"], body.get("warnings"))
        self.assertFalse(body["override"]["capped"])
        self.assertIn("GLOB", {r["ticker"] for r in body["override"]["reduced"]})
        self.assertEqual(self._confirm(body["session_id"]).status_code, 200)
        self.assertAlmostEqual(self._held_in(SIB, "GLOB"), 60, places=3)   # reducido en el sibling
        self.assertAlmostEqual(self._held_in(SIB, "AAPL"), 10, places=3)   # match, intacto

    def _usd_buy(self, broker, tkr, qty, price):
        return NormalizedTx(row_index=0, date="2026-01-10", broker=broker, operation_type=OP_BUY,
                            asset_symbol=tkr, asset_type="CEDEAR", quantity=qty, unit_price=price,
                            gross_amount=qty * price, currency="USD")

    def _preview(self, wb_bytes):
        return self.client.post(
            "/api/imports/tenencia/preview",
            headers={"Authorization": f"Bearer {self.token}"},
            data={"broker": self.BROKER, "format": "ieb"},
            files={"file": ("Portafolio.xlsx", wb_bytes,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})

    def _confirm(self, sid):
        return self.client.post("/api/imports/confirm",
                               headers={"Authorization": f"Bearer {self.token}"},
                               json={"session_id": sid})

    def test_override_reduces_when_not_capped(self):
        """3 activos import-linked; la foto reduce GGAL 100→60 (no capea: 1 de 3, cut
        chico). Tras confirm+rebuild GGAL=60 (venta a costo), MELI/AAPL intactos."""
        self._import_mov([self._buy("MELI", 52, 21112), self._buy("GGAL", 100, 7000),
                          self._buy("AAPL", 10, 20000)])
        self.assertEqual(self._held("GGAL"), 100)
        wb = _ieb_wb([("Cedears", [("MELI", 52, 21112), ("GGAL", 60, 7000), ("AAPL", 10, 20000)])])
        pv = self._preview(wb); self.assertEqual(pv.status_code, 200, pv.text)
        body = pv.json()
        self.assertTrue(body["foto_completa"])
        self.assertFalse(body["override"]["capped"])
        self.assertIn("GGAL", {r["ticker"] for r in body["override"]["reduced"]})
        self.assertEqual(self._confirm(body["session_id"]).status_code, 200)
        self.assertAlmostEqual(self._held("GGAL"), 60, places=4)
        self.assertAlmostEqual(self._held("MELI"), 52, places=4)
        self.assertAlmostEqual(self._held("AAPL"), 10, places=4)

    def test_x3_whole_cartera_is_capped(self):
        """Cartera entera ×3 (MELI 156, un solo activo) vs foto 52: reducir 2/3 dispara
        el cap → NO se auto-reduce (seguro). override.capped=True, MELI queda 156."""
        self._import_mov([self._buy("MELI", 156, 21112)])
        wb = _ieb_wb([("Cedears", [("MELI", 52, 21112)])])
        pv = self._preview(wb); self.assertEqual(pv.status_code, 200, pv.text)
        self.assertTrue(pv.json()["override"]["capped"])
        self._confirm(pv.json()["session_id"])
        self.assertAlmostEqual(self._held("MELI"), 156, places=4)

    def test_partial_foto_does_not_remove(self):
        """Foto PARCIAL (hoja Cauciones no leída → warnings): complete=False → NO borra
        el activo ausente (GGAL sigue), aunque la foto no lo liste."""
        self._import_mov([self._buy("MELI", 52, 21112), self._buy("GGAL", 100, 7000)])
        wb = _ieb_wb([("Cedears", [("MELI", 52, 21112)])], extra_sheet="Cauciones")
        pv = self._preview(wb); self.assertEqual(pv.status_code, 200, pv.text)
        body = pv.json()
        self.assertFalse(body["foto_completa"])
        self.assertEqual(body["override"]["removed"], [])   # no borra por foto parcial
        self._confirm(body["session_id"])
        self.assertAlmostEqual(self._held("GGAL"), 100, places=4)   # sigue tenido


if __name__ == "__main__":
    unittest.main()
