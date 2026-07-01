"""Guards de completitud del parser del Portafolio de IEB (parse_ieb_portfolio).

Estos gatean el modo OVERRIDE en el endpoint: si el parser deja warnings (lectura
parcial), la foto NO borra por 'ausencia'. Acá se testea el PARSER (que los warnings
salgan cuando corresponde y NO en el happy path).

Corre con: cd backend && python3 -m pytest tests/test_ieb_completeness.py
"""
import os
import sys
import unittest
from datetime import datetime

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
BACKEND = os.path.dirname(HERE)
if BACKEND not in sys.path:
    sys.path.insert(0, BACKEND)

from importing.tenencia import parse_ieb_portfolio  # noqa: E402

HDR = ["Especie", "Moneda de emisión", "Cantidad", "Precio", "% del total",
       "PPP", "Var%", "Resultado", "Actualizado", "Posición total"]


def _cedear_row(tkr, qty, ppp="21,112.77"):
    return [f"{tkr} - CEDEAR {tkr}", "ARS", str(qty), "22,680.00", "3.60",
            ppp, "7.42", "81,496.19", "10:43", "1,179,360.00"]


def _stock_row(tkr, qty, ppp="6500"):
    return [f"{tkr} - {tkr} S.A.", "ARS", str(qty), "7000", "5.0",
            ppp, "1.0", "0", "10:00", str(qty * 7000)]


def _wb(sections=None, saldos=True, cauciones=False):
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Patrimonio"
    ws.append(["Fecha:", datetime(2026, 7, 1, 3, 0, 0)])
    ws.append(["Patrimonio total"])
    ws.append(["Tenencia Mercado Argentino"])
    for name, rows in (sections or [("Cedears", [_cedear_row("MELI", 52)])]):
        ws.append([name]); ws.append(list(HDR))
        for r in rows:
            ws.append(r)
        ws.append(["Subtotal", "-", "-", None, "-", "-", "0", "-", "0", None])
    sal = wb.create_sheet("Saldos")
    if saldos:
        sal.append(["ARS"]); sal.append(["Plazo", "Fecha", "Saldo"])
        sal.append(["Hoy", "2026-07-01", "1000"]); sal.append(["Total", "-", "1000"])
        sal.append([]); sal.append(["USD"]); sal.append(["Plazo", "Fecha", "Saldo"])
        sal.append(["Hoy", "2026-07-01", "0"]); sal.append(["Total", "-", "0"])
    else:
        sal.append(["ARS"]); sal.append(["Plazo", "Fecha", "Saldo"])   # sin 'Total'
    if cauciones:
        cau = wb.create_sheet("Cauciones")
        cau.append(["Especie", "Moneda", "Cantidad", "Tasa", "Vencimiento"])
        cau.append(["CAUCION 7D", "ARS", "500000", "40.0", "2026-07-08"])
    return wb


class IebCompletenessTest(unittest.TestCase):
    def test_happy_path_no_warnings(self):
        snap = parse_ieb_portfolio(_wb())
        self.assertEqual(snap.warnings, [])
        self.assertEqual([h.ticker for h in snap.holdings], ["MELI"])
        self.assertAlmostEqual(snap.cash_ars, 1000.0)

    def test_unknown_section_warns_no_contamination(self):
        snap = parse_ieb_portfolio(_wb(sections=[
            ("Cedears", [_cedear_row("MELI", 52)]),
            ("Renta Fija", [["AL30 - BONO", "ARS", "1000", "70.00", "5.0",
                             "65.00", "1.0", "0", "10:00", "70000"]]),
        ]))
        self.assertTrue(snap.warnings)
        self.assertIn("MELI", [h.ticker for h in snap.holdings])
        self.assertNotIn("AL30", [h.ticker for h in snap.holdings])

    def test_multiblock_second_header_synonym_is_read(self):
        HDR2 = ["Ticker", "Moneda de emisión", "Cantidad", "Precio", "% del total",
                "PPP", "Var%", "Resultado", "Actualizado", "Posición total"]
        snap = parse_ieb_portfolio(_wb(sections=[
            ("Acciones", [_stock_row("GGAL", 100)]),
            ("Acciones", [_stock_row("YPFD", 50)]),  # se le pone HDR2 abajo
        ]))
        # Reconstruyo con el 2do header sinónimo:
        wb = _wb(sections=[("Acciones", [_stock_row("GGAL", 100)])])
        ws = wb["Patrimonio"]
        ws.append(["Acciones"]); ws.append(HDR2); ws.append(_stock_row("YPFD", 50))
        ws.append(["Subtotal", "-", "-", None, "-", "-", "0", "-", "0", None])
        snap = parse_ieb_portfolio(wb)
        tickers = [h.ticker for h in snap.holdings]
        self.assertIn("GGAL", tickers)
        self.assertIn("YPFD", tickers)   # header 'Ticker' tolerado → no se pierde

    def test_section_unreadable_header_warns(self):
        wb = _wb(sections=[("Acciones", [_stock_row("GGAL", 100)])])
        ws = wb["Patrimonio"]
        ws.append(["Acciones"]); ws.append(["Columna1", "Columna2", "Columna3"])
        ws.append(_stock_row("YPFD", 50))
        ws.append(["Subtotal", "-", "-", None, "-", "-", "0", "-", "0", None])
        snap = parse_ieb_portfolio(wb)
        self.assertIn("GGAL", [h.ticker for h in snap.holdings])
        self.assertNotIn("YPFD", [h.ticker for h in snap.holdings])
        self.assertTrue(snap.warnings)

    def test_otros_non_dolarusa_warns(self):
        snap = parse_ieb_portfolio(_wb(sections=[
            ("Cedears", [_cedear_row("MELI", 52)]),
            ("Otros", [["SPY - CEDEAR SPY", "ARS", "10", "20000", "0.1",
                        "19000", "1", "0", "10:00", "200000"]]),
        ]))
        self.assertTrue(any("Otros" in w for w in snap.warnings))
        self.assertNotIn("SPY", [h.ticker for h in snap.holdings])

    def test_cauciones_sheet_warns(self):
        snap = parse_ieb_portfolio(_wb(cauciones=True))
        self.assertTrue(any("Cauciones" in w for w in snap.warnings))

    def test_missing_saldos_warns(self):
        snap = parse_ieb_portfolio(_wb(saldos=False))
        self.assertTrue(any("saldos" in w.lower() for w in snap.warnings))


if __name__ == "__main__":
    unittest.main()
