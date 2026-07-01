"""Parser del Portafolio (tenencia) de IEB — importing/tenencia.parse_ieb_portfolio.

El export de IEB (Portafolio ARS / Portafolio USD) es un Excel con hojas
Patrimonio (tenencias, con PPP=costo) + Saldos (cash). Es la FOTO de posiciones
de HOY: siembra el costo real (no P&L 0) y pisa lo que el historial dejó.

Corre con: cd backend && python3 -m pytest tests/test_ieb_portfolio.py
"""
import os
import sys
import unittest
from datetime import datetime

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
BACKEND = os.path.dirname(HERE)
if BACKEND not in sys.path:
    sys.path.insert(0, BACKEND)

from importing.tenencia import parse_ieb_portfolio, looks_like_ieb_portfolio  # noqa: E402

HDR = ["Especie", "Moneda de emisión", "Cantidad", "Precio", "% del total",
       "PPP", "Var%", "Resultado", "Actualizado", "Posición total"]


def _wb():
    """Reproduce la estructura real (números en formato US: coma miles)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Patrimonio"
    ws.append(["Fecha:", datetime(2026, 7, 1, 3, 0, 0)])
    ws.append(["Patrimonio total"])
    ws.append(["Tenencia Mercado Argentino"])
    ws.append(["Cedears"])
    ws.append(HDR)
    ws.append(["MELI - CEDEAR MERCADOLIBRE INC", "ARS", "52", "22,680.00", "3.60",
               "21,112.77", "7.42", "81,496.19", "10:43", "1,179,360.00"])
    ws.append(["Disponible", "-", "52", "22,680.00", "-", "-", "-", "-", "-", "1,179,360.00"])
    ws.append(["Subtotal", "-", "-", None, "-", "-", "81,496.19", "-", "1,179,360.00", None])
    ws.append(["Otros"])
    ws.append(HDR)
    ws.append(["DOLARUSA - DOLARES USA ESP 7000", "USD", "2", "1,539.00", "0.01",
               "1,490.66", "3.24", "104.41", "10:42", "3,324.24"])
    ws.append(["Subtotal", "-", "-", "0.01", "-", "-", "104.41", "-", "3,324.24", None])
    sal = wb.create_sheet("Saldos")
    sal.append(["ARS"]); sal.append(["Plazo", "Fecha", "Saldo"])
    sal.append(["Hoy", "2026-07-01", "1000"]); sal.append(["Total", "-", "1000"])
    sal.append([]); sal.append([])
    sal.append(["USD"]); sal.append(["Plazo", "Fecha", "Saldo"])
    sal.append(["Hoy", "2026-07-01", "97.29"]); sal.append(["Total", "-", "97.29"])
    return wb


class IebPortfolioParseTest(unittest.TestCase):
    def test_detects_format(self):
        self.assertTrue(looks_like_ieb_portfolio(_wb()))
        wb = openpyxl.Workbook()  # sin las hojas → no matchea
        self.assertFalse(looks_like_ieb_portfolio(wb))

    def test_holdings_cost_and_cash(self):
        snap = parse_ieb_portfolio(_wb())
        # 1 holding real (MELI); DOLARUSA es cash, NO holding.
        self.assertEqual(len(snap.holdings), 1)
        h = snap.holdings[0]
        self.assertEqual(h.ticker, "MELI")
        self.assertEqual(h.asset_type, "CEDEAR")
        self.assertAlmostEqual(h.quantity, 52.0)
        self.assertEqual(h.currency, "ARS")
        # el PPP (costo promedio) se usa como precio de siembra → P&L real
        self.assertAlmostEqual(h.price_per1, 21112.77, places=2)
        # cash de la hoja Saldos
        self.assertAlmostEqual(snap.cash_ars, 1000.0)
        self.assertAlmostEqual(snap.cash_usd, 97.29)
        self.assertEqual(snap.date, "2026-07-01")

    def test_dolarusa_not_a_holding(self):
        snap = parse_ieb_portfolio(_wb())
        self.assertNotIn("DOLARUSA", [h.ticker for h in snap.holdings])


if __name__ == "__main__":
    unittest.main()
